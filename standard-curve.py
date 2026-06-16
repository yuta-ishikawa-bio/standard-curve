"""検量線自動作成ツール (Calibration Curve Generator)

CSV / Excel の測定データから最小二乗法で検量線 (直線 y = ax + b) を作成し、
回帰式・決定係数 (R^2)・散布図 (近似直線つき) を含む Excel レポートを出力する。
未知サンプルの測定値から濃度を逆算する機能も備える。

使い方:
    # ファイルから読み込み (CSV または Excel)
    python standard-curve.py --input sample_data/example.csv

    # 出力ファイル名を指定
    python standard-curve.py --input data.xlsx --output result.xlsx

    # 未知サンプルの測定値から濃度を逆算 (複数指定可)
    python standard-curve.py --input data.csv --unknown 0.234 0.512

    # コマンドラインで手入力
    python standard-curve.py --manual
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from scipy import stats

from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.trendline import Trendline
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass
class CalibrationResult:
    """検量線の回帰結果を保持するデータクラス。"""

    slope: float        # 傾き a
    intercept: float    # 切片 b
    r_squared: float    # 決定係数 R^2
    n_points: int       # 使用したデータ点数

    @property
    def equation(self) -> str:
        sign = "+" if self.intercept >= 0 else "-"
        return f"y = {self.slope:.6g}x {sign} {abs(self.intercept):.6g}"

    def estimate_concentration(self, signal: float) -> float:
        """測定値 y から濃度 x を逆算する。"""
        if self.slope == 0:
            raise ValueError("傾きが 0 のため濃度を逆算できません。")
        return (signal - self.intercept) / self.slope


def load_data(path: Path) -> pd.DataFrame:
    """CSV または Excel から測定データを読み込む。

    先頭 2 列をそれぞれ「濃度 (x)」「測定値 (y)」として解釈する。
    """
    if not path.exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {path}")

    suffix = path.suffix.lower()
    if suffix == ".csv":
        df = pd.read_csv(path)
    elif suffix in (".xlsx", ".xlsm"):
        df = pd.read_excel(path)
    else:
        raise ValueError(f"対応していない拡張子です: {suffix} (.csv / .xlsx)")

    if df.shape[1] < 2:
        raise ValueError("データには最低 2 列 (濃度・測定値) が必要です。")

    df = df.iloc[:, :2].copy()
    df.columns = ["concentration", "signal"]
    df = df.apply(pd.to_numeric, errors="coerce").dropna()

    if len(df) < 2:
        raise ValueError("回帰には有効な数値データが 2 点以上必要です。")

    return df.reset_index(drop=True)


def input_data_manually() -> pd.DataFrame:
    """コマンドラインから濃度・測定値を対話的に入力する。"""
    print("濃度と測定値をスペース区切りで入力してください (空行で終了)")
    print("例: 0 0.012")
    rows: list[tuple[float, float]] = []
    while True:
        try:
            line = input(f"[{len(rows) + 1}] > ").strip()
        except EOFError:
            break
        if not line:
            break
        parts = line.split()
        if len(parts) != 2:
            print("  -> 2 個の数値を入力してください。")
            continue
        try:
            x, y = float(parts[0]), float(parts[1])
        except ValueError:
            print("  -> 数値として解釈できません。")
            continue
        rows.append((x, y))

    if len(rows) < 2:
        raise ValueError("回帰には 2 点以上のデータが必要です。")

    return pd.DataFrame(rows, columns=["concentration", "signal"])


def fit_calibration(df: pd.DataFrame) -> CalibrationResult:
    """最小二乗法で直線回帰を行い、検量線を求める。"""
    result = stats.linregress(df["concentration"], df["signal"])
    return CalibrationResult(
        slope=float(result.slope),
        intercept=float(result.intercept),
        r_squared=float(result.rvalue) ** 2,
        n_points=len(df),
    )


def write_report(
    df: pd.DataFrame,
    result: CalibrationResult,
    output_path: Path,
    unknowns: list[float] | None = None,
) -> None:
    """検量線の結果を Excel レポートとして出力する。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "検量線"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    center = Alignment(horizontal="center")

    # --- 測定データ表 ---
    ws["A1"] = "濃度 (x)"
    ws["B1"] = "測定値 (y)"
    for cell in (ws["A1"], ws["B1"]):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for i, row in df.iterrows():
        ws.cell(row=i + 2, column=1, value=float(row["concentration"]))
        ws.cell(row=i + 2, column=2, value=float(row["signal"]))

    last_data_row = len(df) + 1

    # --- 回帰結果ブロック ---
    summary_start = last_data_row + 2
    summary = [
        ("回帰式", result.equation),
        ("傾き a", result.slope),
        ("切片 b", result.intercept),
        ("決定係数 R^2", result.r_squared),
        ("データ点数", result.n_points),
    ]
    ws.cell(row=summary_start, column=1, value="回帰結果").font = Font(bold=True)
    for j, (label, value) in enumerate(summary):
        r = summary_start + 1 + j
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)

    # --- 散布図 + 近似直線 ---
    chart = ScatterChart()
    chart.title = "検量線 (Calibration Curve)"
    chart.x_axis.title = "濃度 (x)"
    chart.y_axis.title = "測定値 (y)"
    chart.scatterStyle = "marker"
    chart.height = 9
    chart.width = 15

    x_ref = Reference(ws, min_col=1, min_row=2, max_row=last_data_row)
    y_ref = Reference(ws, min_col=2, min_row=2, max_row=last_data_row)
    series = Series(y_ref, x_ref, title="測定データ")

    # マーカーのみ表示し、点をつなぐ線は消す
    series.marker = Marker(symbol="circle", size=7)
    series.graphicalProperties = GraphicalProperties()
    series.graphicalProperties.line.noFill = True

    # 近似直線 (回帰式と R^2 をグラフ上に表示)
    series.trendline = Trendline(
        trendlineType="linear", dispEq=True, dispRSqr=True
    )

    chart.series.append(series)
    ws.add_chart(chart, f"{get_column_letter(4)}2")  # D2 に配置

    # --- 未知サンプルの濃度逆算 ---
    if unknowns:
        u_start = summary_start + len(summary) + 3
        ws.cell(row=u_start, column=1, value="未知サンプルの濃度逆算").font = Font(bold=True)
        h1 = ws.cell(row=u_start + 1, column=1, value="測定値 (y)")
        h2 = ws.cell(row=u_start + 1, column=2, value="推定濃度 (x)")
        for cell in (h1, h2):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        for k, y in enumerate(unknowns):
            ws.cell(row=u_start + 2 + k, column=1, value=float(y))
            ws.cell(row=u_start + 2 + k, column=2, value=result.estimate_concentration(y))

    # 列幅調整
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16

    wb.save(output_path)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="CSV/Excel の測定データから検量線 (直線) を作成し Excel に出力する。"
    )
    parser.add_argument("--input", "-i", type=Path, help="入力ファイル (.csv / .xlsx)")
    parser.add_argument("--manual", "-m", action="store_true", help="コマンドラインで手入力する")
    parser.add_argument(
        "--output", "-o", type=Path, default=Path("calibration_result.xlsx"),
        help="出力 Excel ファイル名 (既定: calibration_result.xlsx)",
    )
    parser.add_argument(
        "--unknown", "-u", type=float, nargs="*", default=None,
        help="未知サンプルの測定値 (濃度を逆算する。複数指定可)",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)

    if args.manual:
        df = input_data_manually()
    elif args.input:
        df = load_data(args.input)
    else:
        print("エラー: --input でファイルを指定するか --manual を使ってください。", file=sys.stderr)
        return 1

    result = fit_calibration(df)

    print("=== 検量線 ===")
    print(f"  回帰式      : {result.equation}")
    print(f"  決定係数 R^2: {result.r_squared:.6f}")
    print(f"  データ点数  : {result.n_points}")

    if args.unknown:
        print("--- 未知サンプルの濃度逆算 ---")
        for y in args.unknown:
            print(f"  測定値 {y:g} -> 推定濃度 {result.estimate_concentration(y):.6g}")

    write_report(df, result, args.output, args.unknown)
    print(f"Excel レポートを出力しました: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())